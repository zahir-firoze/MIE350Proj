import openpyxl
import numpy as np
import random as rd
import string

from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active

upc_nums = ['45496590741',
'45496590420',
'45496593858',
'45496596484',
'4549659655',
'0045496422899',
'21112248920',
'45496901998',
'45496742805',
'45496741747',
'045496743888',
'730865020102',
'810023033219',
'014633738179',
'662248915050',
'71042559529',
'730865200009',
'083717202394',
'014633354522',
'711719221524',
'696554197565',
'113102814465',
'711719504405',
'045496891947',
'711719220381']

emails=['janneh@verizon.net',
'benits@me.com',
'jonathan@hotmail.com',
'claypool@gmail.com',
'pereinar@outlook.com',
'adamk@live.com',
'shang@msn.com',
'emcleod@aol.com',
'harpes@live.com',
'dmiller@outlook.com',
'wildfire@sbcglobal.net',
'jbryan@msn.com',
    'bowmanbs@me.com',
    'evans@comcast.net',
    'debest@me.com',
    'jandrese@gmail.com',
    'kosact@gmail.com',
    'djupedal@gmail.com',
    'jpflip@me.com',
    'osaru@yahoo.com',
    'neuffer@sbcglobal.net',
    'lauronen@comcast.net',
    'msherr@comcast.net',
    'dgatwood@yahoo.ca',
        'mccurley@mac.com',
    'wayward@hotmail.com',
    'penna@optonline.net',
    'metzzo@comcast.net',
    'mchugh@msn.com',
    'nogin@mac.com',
    'esasaki@aol.com',
    'parksh@yahoo.ca',
    'maikelnai@mac.com',
    'leslie@verizon.net',
    'purvis@hotmail.com',
    'empathy@verizon.net',
        'parsimony@outlook.com',
    'mmccool@aol.com',
    'garyjb@sbcglobal.net',
    'fviegas@aol.com',
    'smpeters@me.com',
    'wildfire@yahoo.ca',
    'agolomsh@att.net',
    'draper@me.com',
    'adillon@sbcglobal.net',
    'hllam@hotmail.com',
    'floxy@icloud.com',
    'scotfl@icloud.com',
        'ismail@hotmail.com',
    'budinger@live.com',
    'gozer@icloud.com',
    'jaxweb@att.net',
    'monopole@msn.com',
    'janusfury@verizon.net',
    'pspoole@outlook.com',
    'ghost@optonline.net',
    'sumdumass@yahoo.ca',
    'nwiger@yahoo.com',
    'helger@outlook.com',
    'andrei@optonline.net',
        'mchugh@msn.com',
    'salesgeek@hotmail.com',
    'jschauma@verizon.net',
    'mcmillan@me.com',
    'dbrobins@me.com',
    'ivoibs@optonline.net',
    'earmstro@att.net',
    'vsprintf@hotmail.com',
    'luebke@outlook.com',
    'neuffer@outlook.com',
    'heine@aol.com',
    'nimaclea@msn.com',
        'wetter@icloud.com',
    'world@att.net',
    'techie@comcast.net',
    ]

review = ['Worst game ever!', 'Not fun.', 'It was average.', 'Nice game!', 'Best game ever!']

count = 1
for upc in upc_nums:
    ws['A{0}'.format(count)] = upc
    ws['A{0}'.format(count+1)] = upc
    ws['A{0}'.format(count+2)] = upc

    ws['B{0}'.format(count)] = emails[count-1]
    ws['B{0}'.format(count+1)] = emails[count]
    ws['B{0}'.format(count+2)] = emails[count+1]

    ws['C{0}'.format(count)] = np.random.randint(1,6)
    ws['C{0}'.format(count+1)] = np.random.randint(1,6)
    ws['C{0}'.format(count+2)] = np.random.randint(1,6)

    if (ws['C{0}'.format(count)].value == 1):
        ws['D{0}'.format(count)] = review[0]
    elif(ws['C{0}'.format(count)].value == 2):
        ws['D{0}'.format(count)] = review[1]
    elif(ws['C{0}'.format(count)].value == 3):
        ws['D{0}'.format(count)] = review[2]
    elif(ws['C{0}'.format(count)].value == 4):
        ws['D{0}'.format(count)] = review[3]
    elif(ws['C{0}'.format(count)].value == 5):
        ws['D{0}'.format(count)] = review[4]

    if (ws['C{0}'.format(count+1)].value == 1):
        ws['D{0}'.format(count+1)] = review[0]
    elif(ws['C{0}'.format(count+1)].value == 2):
        ws['D{0}'.format(count+1)] = review[1]
    elif(ws['C{0}'.format(count+1)].value == 3):
        ws['D{0}'.format(count+1)] = review[2]
    elif(ws['C{0}'.format(count+1)].value == 4):
        ws['D{0}'.format(count+1)] = review[3]
    elif(ws['C{0}'.format(count+1)].value == 5):
        ws['D{0}'.format(count+1)] = review[4]  

    if (ws['C{0}'.format(count+2)].value == 1):
        ws['D{0}'.format(count+2)] = review[0]
    elif(ws['C{0}'.format(count+2)].value == 2):
        ws['D{0}'.format(count+2)] = review[1]
    elif(ws['C{0}'.format(count+2)].value == 3):
        ws['D{0}'.format(count+2)] = review[2]
    elif(ws['C{0}'.format(count+2)].value == 4):
        ws['D{0}'.format(count+2)] = review[3]
    elif(ws['C{0}'.format(count+2)].value == 5):
        ws['D{0}'.format(count+2)] = review[4]        

    count = count + 3

wb.save('union_of_matthews2.xlsx')
